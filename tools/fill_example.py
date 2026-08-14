"""Populate a copy of the workbook with the worked example, for demonstration
and as a cross-check that the Excel formulas agree with the Python engine.

    python tools/fill_example.py
"""
from __future__ import annotations

import random
import shutil
import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
from tidcmm.model import load_model  # noqa: E402
from tools.build_workbook import _clean_office_metadata  # noqa: E402

# Version comes from the model so the filenames cannot drift from it.
import yaml as _yaml
FILEV = ".".join(
    _yaml.safe_load((ROOT / "model" / "meta.yaml").read_text())["model"]["version"].split(".")[:2])
SRC = ROOT / "build" / f"TID-CMM-Self-Assessment-v{FILEV}.xlsx"
DST = ROOT / "build" / f"TID-CMM-Worked-Example-v{FILEV}.xlsx"


def main() -> Path:
    model = load_model()
    data = yaml.safe_load((ROOT / "assessments" / "example-assessment.yaml").read_text())
    shutil.copy(SRC, DST)
    wb = load_workbook(DST)

    st = wb["Setup"]
    st["C4"] = data["organisation"]
    st["C5"] = data["assessed_on"]
    st["C6"] = data["assessor"]
    st["C7"] = data["scope"]
    st["C8"] = "Recently acquired payments subsidiary; OT estate at two manufacturing sites."
    st["C9"] = "TP-2026-Q2 (ransomware affiliates, financially motivated intrusion sets, state-aligned espionage)"
    st["C10"] = "Level 3.5 overall by Q4 2027, with AV at Level 4"
    st["C11"] = "2025-09-15"
    st["C12"] = 2.05
    st["C13"] = "Yes"

    for d in model.domains:
        ws = wb[d.id]
        for i, sc in enumerate(d.subcapabilities):
            row = 6 + i
            r = data["responses"][sc.id]
            ws.cell(row=row, column=5, value=r["score"])
            ws.cell(row=row, column=6, value=r["target"])
            ws.cell(row=row, column=8, value=r.get("evidence", ""))
            ws.cell(row=row, column=9, value=r.get("owner", ""))

    # A representative in-scope set with a plausible coverage distribution.
    cov = wb["ATT&CK Coverage"]
    rng = random.Random(20260810)
    in_scope_platforms = {"Windows", "Linux", "macOS", "IaaS", "SaaS", "Identity Provider", "Office Suite", "Containers"}
    counts = {0: 0, 1: 0, 2: 0, 3: 0}
    n_scope = 0
    for row in range(5, cov.max_row + 1):
        tid = cov.cell(row=row, column=1).value
        if not tid:
            continue
        plats = {p.strip() for p in str(cov.cell(row=row, column=5).value or "").split(";")}
        if not (plats & in_scope_platforms):
            cov.cell(row=row, column=6, value="N")
            continue
        cov.cell(row=row, column=6, value="Y")
        n_scope += 1
        status = rng.choices([0, 1, 2, 3], weights=[18, 34, 33, 15])[0]
        cov.cell(row=row, column=7, value=status)
        counts[status] += 1
        if status == 3:
            cov.cell(row=row, column=8, value=rng.choice(
                ["PT-2026-Q2 run 14", "Purple exercise 2026-05", "ART quarterly 2026-Q2", "BAS scenario 118"]))
    wb.properties.creator = "Reza Adineh"
    wb.properties.lastModifiedBy = "Reza Adineh"
    wb.properties.title = "TID-CMM Worked Example"
    wb.save(DST)
    _clean_office_metadata(DST)
    print(f"{DST}  in-scope={n_scope} distribution={counts}")
    return DST


if __name__ == "__main__":
    main()
