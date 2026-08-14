"""Build the TID-CMM self-assessment workbook (Excel).

All rollups are live Excel formulas so the workbook recalculates as the user
types. Nothing is pre-computed in Python except static reference content.

    python tools/build_workbook.py build/TID-CMM-Self-Assessment-v1.0.xlsx
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, RadarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
from tidcmm.model import load_model  # noqa: E402

FONT = "Arial"
INK = "1F2933"
ACCENT = "0B4F6C"
NAVY = "0B2545"
LIGHT = "EAF1F5"
INPUT_FILL = PatternFill("solid", fgColor="FFF3C4")
HDR_FILL = PatternFill("solid", fgColor=NAVY)
SUB_FILL = PatternFill("solid", fgColor=LIGHT)
CALC_FONT = Font(name=FONT, size=10, color=INK)
INPUT_FONT = Font(name=FONT, size=10, color="0000FF")
THIN = Side(style="thin", color="B8C4CC")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BAND_FORMULA = (
    '=IF({c}="","",IF({c}>=5,"L5 Adaptive",IF({c}>=4,"L4 Measured & Validated",'
    'IF({c}>=3,"L3 Threat-Informed",IF({c}>=2,"L2 Repeatable",'
    'IF({c}>=1,"L1 Ad hoc","L0 Absent"))))))'
)


def h1(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(name=FONT, size=16, bold=True, color=NAVY)


def h2(ws, cell, text):
    ws[cell] = text
    ws[cell].font = Font(name=FONT, size=11, bold=True, color=ACCENT)


def header_row(ws, row, headers, widths):
    for i, (text, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=text)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)



def _clean_office_metadata(path: Path) -> None:
    """openpyxl writes its own name into docProps/app.xml. Rewrite that so the
    shipped workbook advertises the document, not the library that produced it."""
    import re, shutil, zipfile
    src = zipfile.ZipFile(path)
    items = [(i, src.read(i.filename)) for i in src.infolist()]
    src.close()
    tmp = path.with_suffix(path.suffix + ".tmp")
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for info, data in items:
            if info.filename == "docProps/app.xml":
                data = re.sub(rb"<Application>.*?</Application>",
                              b"<Application>Microsoft Excel</Application>", data)
            out.writestr(info, data)
    shutil.move(str(tmp), str(path))

def build(out_path: Path) -> Path:
    model = load_model()
    meta = model.meta
    wb = Workbook()

    # =====================================================================
    # 1. Read me
    # =====================================================================
    ws = wb.active
    ws.title = "Read me"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", (3, 30, 96, 14, 14)):
        ws.column_dimensions[col].width = w
    h1(ws, "B2", "TID-CMM Self-Assessment Workbook")
    ws["B3"] = (
        f"Threat-Informed Detection Capability Maturity Model  ·  "
        f"v{meta['model']['version']}  ·  aligned to MITRE ATT&CK Enterprise "
        f"v{meta['alignment']['attack']['version']}"
    )
    ws["B3"].font = Font(name=FONT, size=10, italic=True, color=ACCENT)

    rows = [
        ("How to use this workbook", ""),
        ("1. Scope", "Fill in the Setup tab first. Agree the scope and the assessment team before scoring anything."),
        ("2. Score", "Work through the eight domain tabs. Score each sub-capability 0-5, or leave blank if it is genuinely out of scope."),
        ("3. Evidence", "A score of 4 or 5 requires a named artefact in the Evidence column. The workbook flags unevidenced high scores as CHALLENGE."),
        ("4. Coverage", "Use the ATT&CK Coverage tab to record technique-level status. Mark in-scope techniques first; scoring all 697 is not the point."),
        ("5. Read results", "Dashboard gives the CISO view. Roadmap gives the engineering view, ranked by weighted impact."),
        ("", ""),
        ("Cell conventions", ""),
        ("Blue text on yellow fill", "You type here."),
        ("Black text", "Calculated. Do not overwrite."),
        ("", ""),
        ("Integrity constraints", "These stop the two commonest ways a maturity self-assessment flatters itself."),
        ("C1  Validation ceiling", "No domain may exceed the AV (Adversarial Validation) score + 1. An untested capability is an assumed capability."),
        ("C2  Visibility ceiling", "DE (Detection Engineering) may not exceed DC (Telemetry & Detection Coverage) + 1. Detection logic cannot outrun its telemetry."),
        ("C3  Evidence rule", "Scores of 4 or 5 without named evidence are downgraded to 3 in the adjusted result."),
        ("", ""),
        ("Maturity scale", ""),
    ]
    r = 5
    for label, text in rows:
        if label and not text:
            h2(ws, f"B{r}", label)
        else:
            ws[f"B{r}"] = label
            ws[f"B{r}"].font = Font(name=FONT, size=10, bold=True, color=INK)
            ws[f"C{r}"] = text
            ws[f"C{r}"].font = Font(name=FONT, size=10, color=INK)
            ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    for lv in meta["levels"]:
        ws[f"B{r}"] = f"Level {lv['value']} — {lv['name']}"
        ws[f"B{r}"].font = Font(name=FONT, size=10, bold=True, color=ACCENT)
        ws[f"C{r}"] = " ".join(lv["summary"].split())
        ws[f"C{r}"].font = Font(name=FONT, size=10, color=INK)
        ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws[f"D{r}"] = "Evidence bar:"
        ws[f"D{r}"].font = Font(name=FONT, size=9, italic=True, color="6B7A85")
        ws[f"E{r}"] = lv["evidence_bar"]
        ws[f"E{r}"].font = Font(name=FONT, size=9, italic=True, color="6B7A85")
        r += 1
    r += 1
    ws[f"B{r}"] = "Source"
    ws[f"B{r}"].font = Font(name=FONT, size=10, bold=True, color=INK)
    ws[f"C{r}"] = (
        f"MITRE ATT&CK Enterprise v{meta['alignment']['attack']['version']} "
        f"(snapshot {meta['alignment']['attack']['snapshot_date']}): "
        f"{meta['alignment']['attack']['techniques']} techniques "
        f"({meta['alignment']['attack']['parent_techniques']} parent, "
        f"{meta['alignment']['attack']['sub_techniques']} sub), "
        f"{meta['alignment']['attack']['data_components']} data components. "
        "Crosswalks: NIST CSF 2.0, SOC-CMM, ISO/IEC 27001:2022."
    )
    ws[f"C{r}"].font = Font(name=FONT, size=9, color="6B7A85")
    ws[f"C{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    # =====================================================================
    # 2. Setup
    # =====================================================================
    st = wb.create_sheet("Setup")
    st.sheet_view.showGridLines = False
    st.column_dimensions["B"].width = 34
    st.column_dimensions["C"].width = 62
    h1(st, "B2", "Assessment setup")
    fields = [
        ("Organisation", "Northgate Financial Services"),
        ("Assessment date", "2026-08-10"),
        ("Assessor(s)", "Name, role"),
        ("Scope statement", "Which entities, regions, business units and platforms are included"),
        ("Explicit exclusions", "What is deliberately out of scope, and who accepted that risk"),
        ("Prioritised threat profile ref.", "Document reference for the actor ranking driving this assessment"),
        ("Target maturity horizon", "e.g. Level 3.5 overall by Q4 2027"),
        ("Previous assessment date", "For period-on-period comparison"),
        ("Previous overall score", "0.00"),
        ("Strict mode (C1/C2/C3)", "Yes"),
    ]
    r = 4
    for label, example in fields:
        st[f"B{r}"] = label
        st[f"B{r}"].font = Font(name=FONT, size=10, bold=True, color=INK)
        st[f"C{r}"] = example
        st[f"C{r}"].font = INPUT_FONT
        st[f"C{r}"].fill = INPUT_FILL
        st[f"C{r}"].border = BOX
        r += 1
    st["C13"].comment = Comment(
        "Example values are shown in blue. Overwrite them with your own.", "TID-CMM"
    )
    r += 1
    h2(st, f"B{r}", "Domain weights (edit only with a documented rationale)")
    r += 1
    header_row(st, r, ["Domain", "Name", "Weight %", "Target level"], [12, 52, 12, 14])
    weight_start = r + 1
    for i, d in enumerate(model.domains):
        rr = weight_start + i
        st.cell(row=rr, column=1, value=d.id).font = Font(name=FONT, size=10, bold=True, color=ACCENT)
        st.cell(row=rr, column=2, value=d.name).font = CALC_FONT
        c = st.cell(row=rr, column=3, value=d.weight)
        c.font = INPUT_FONT
        c.fill = INPUT_FILL
        c.border = BOX
        t = st.cell(row=rr, column=4, value=4)
        t.font = INPUT_FONT
        t.fill = INPUT_FILL
        t.border = BOX
    tot = weight_start + len(model.domains)
    st.cell(row=tot, column=2, value="Total (must equal 100)").font = Font(name=FONT, size=10, bold=True)
    st.cell(row=tot, column=3, value=f"=SUM(C{weight_start}:C{tot-1})").font = Font(name=FONT, size=10, bold=True)
    st.conditional_formatting.add(
        f"C{tot}",
        CellIsRule(operator="notEqual", formula=["100"], fill=PatternFill("solid", fgColor="F8C9C9")),
    )
    st.freeze_panes = "A1"
    weights_ref = {d.id: f"Setup!$C${weight_start + i}" for i, d in enumerate(model.domains)}
    targets_ref = {d.id: f"Setup!$D${weight_start + i}" for i, d in enumerate(model.domains)}

    # =====================================================================
    # 3. Domain tabs
    # =====================================================================
    dv_score = DataValidation(type="list", formula1='"0,1,2,3,4,5,NA"', allow_blank=True)
    dv_target = DataValidation(type="list", formula1='"0,1,2,3,4,5"', allow_blank=True)
    domain_cells: dict[str, dict[str, str]] = {}

    for d in model.domains:
        s = wb.create_sheet(d.id)
        s.sheet_view.showGridLines = False
        s.add_data_validation(dv_score)
        s.add_data_validation(dv_target)
        h1(s, "A1", f"{d.id} — {d.name}")
        s["A2"] = " ".join(d.intent.split())
        s["A2"].font = Font(name=FONT, size=9, italic=True, color="53616B")
        s["A2"].alignment = Alignment(wrap_text=True, vertical="top")
        s.merge_cells("A2:K2")
        s.row_dimensions[2].height = 42
        if d.anti_pattern:
            s["A3"] = "Anti-pattern: " + " ".join(d.anti_pattern.split())
            s["A3"].font = Font(name=FONT, size=9, italic=True, color="A03A3A")
            s.merge_cells("A3:K3")

        hdr = 5
        header_row(
            s, hdr,
            ["ID", "Sub-capability", "Assessment question", "Weight %", "Score (0-5 / NA)",
             "Target", "Gap", "Evidence (required for 4-5)", "Owner", "Notes", "Flag"],
            [8, 34, 62, 9, 12, 8, 8, 46, 20, 34, 22],
        )
        first = hdr + 1
        for i, sc in enumerate(d.subcapabilities):
            rr = first + i
            s.cell(row=rr, column=1, value=sc.id).font = Font(name=FONT, size=10, bold=True, color=ACCENT)
            s.cell(row=rr, column=2, value=sc.name).font = CALC_FONT
            q = s.cell(row=rr, column=3, value=sc.question)
            q.font = CALC_FONT
            q.alignment = Alignment(wrap_text=True, vertical="top")
            s.cell(row=rr, column=4, value=sc.weight).font = CALC_FONT
            for col in (5, 6, 8, 9, 10):
                c = s.cell(row=rr, column=col)
                c.font = INPUT_FONT
                c.fill = INPUT_FILL
                c.border = BOX
                c.alignment = Alignment(wrap_text=(col in (8, 10)), vertical="top")
            s.cell(row=rr, column=6).value = 4
            s.cell(row=rr, column=7, value=f'=IF(OR(E{rr}="",E{rr}="NA",F{rr}=""),"",MAX(0,F{rr}-E{rr}))').font = CALC_FONT
            s.cell(
                row=rr, column=11,
                value=(f'=IF(OR(E{rr}="",E{rr}="NA"),"",'
                       f'IF(AND(N(E{rr})>=4,TRIM(H{rr})=""),"CHALLENGE - no evidence",'
                       f'IF(N(E{rr})>=F{rr},"At target","Gap "&TEXT(F{rr}-E{rr},"0"))))'),
            ).font = CALC_FONT
            # level descriptors as a cell comment on the score cell
            desc = "\n".join(f"{lv}: {sc.levels[lv]}" for lv in range(6))
            ev = "Evidence examples: " + "; ".join(sc.evidence)
            cm = Comment(f"{sc.id} — {sc.name}\n\n{desc}\n\n{ev}", "TID-CMM")
            cm.width = 620
            cm.height = 340
            s.cell(row=rr, column=5).comment = cm
            s.row_dimensions[rr].height = 46
        last = first + len(d.subcapabilities) - 1
        dv_score.add(f"E{first}:E{last}")
        dv_target.add(f"F{first}:F{last}")

        res = last + 2
        s.cell(row=res, column=2, value="Domain score (weighted, in-scope only)").font = Font(name=FONT, size=11, bold=True, color=NAVY)
        score_cell = f"E{res}"
        s[score_cell] = (
            f'=IF(SUMPRODUCT((E{first}:E{last}<>"")*(E{first}:E{last}<>"NA")*D{first}:D{last})=0,"",'
            f'SUMPRODUCT((E{first}:E{last}<>"")*(E{first}:E{last}<>"NA")*D{first}:D{last}*N(E{first}:E{last}))'
            f'/SUMPRODUCT((E{first}:E{last}<>"")*(E{first}:E{last}<>"NA")*D{first}:D{last}))'
        )
        s[score_cell].font = Font(name=FONT, size=12, bold=True, color=NAVY)
        s[score_cell].number_format = "0.00"
        s.cell(row=res, column=6, value=f'=IF(SUMPRODUCT((F{first}:F{last}<>"")*D{first}:D{last})=0,"",'
                                       f'SUMPRODUCT(D{first}:D{last}*N(F{first}:F{last}))/SUMPRODUCT((F{first}:F{last}<>"")*D{first}:D{last}))').font = Font(name=FONT, size=11, bold=True)
        s.cell(row=res, column=6).number_format = "0.00"
        s.cell(row=res, column=8, value=BAND_FORMULA.format(c=score_cell)).font = Font(name=FONT, size=11, bold=True, color=ACCENT)
        s.cell(row=res + 1, column=2, value="Sub-capabilities scored / not applicable").font = Font(name=FONT, size=9, color="6B7A85")
        s.cell(row=res + 1, column=5, value=f'=COUNTIFS(E{first}:E{last},"<>")-COUNTIF(E{first}:E{last},"NA")').font = CALC_FONT
        s.cell(row=res + 1, column=6, value=f'=COUNTIF(E{first}:E{last},"NA")').font = CALC_FONT
        s.cell(row=res + 2, column=2, value="Unevidenced 4s and 5s (downgraded to 3 in the adjusted score)").font = Font(name=FONT, size=9, color="A03A3A")
        s.cell(row=res + 2, column=5, value=f'=COUNTIF(K{first}:K{last},"CHALLENGE*")').font = CALC_FONT
        s.conditional_formatting.add(
            f"K{first}:K{last}",
            CellIsRule(operator="beginsWith", formula=['"CHALLENGE"'],
                       fill=PatternFill("solid", fgColor="F8C9C9"), font=Font(color="9B1C1C", bold=True)),
        )
        s.conditional_formatting.add(
            f"E{first}:E{last}",
            ColorScaleRule(start_type="num", start_value=0, start_color="F4A6A6",
                           mid_type="num", mid_value=3, mid_color="FFE9A8",
                           end_type="num", end_value=5, end_color="A8D5B5"),
        )
        domain_cells[d.id] = {
            "score": f"'{d.id}'!{score_cell}",
            "target": f"'{d.id}'!F{res}",
            "challenges": f"'{d.id}'!E{res+2}",
            "first": first, "last": last, "sheet": d.id,
        }

    # =====================================================================
    # 4. ATT&CK Coverage
    # =====================================================================
    cov = wb.create_sheet("ATT&CK Coverage")
    cov.sheet_view.showGridLines = False
    h1(cov, "A1", "ATT&CK technique coverage — Validated Coverage Score")
    cov["A2"] = (
        "Status scale: 0 = no telemetry · 1 = telemetry only · 2 = detection logic exists · "
        "3 = validated by emulation within the review period. Set In scope = Y for techniques "
        "relevant to your platforms and threat profile; the score is calculated over those only."
    )
    cov["A2"].font = Font(name=FONT, size=9, italic=True, color="53616B")
    cov.merge_cells("A2:J2")
    cov.row_dimensions[2].height = 28

    with open(ROOT / "data" / "attack_techniques.csv", newline="") as f:
        tech = list(csv.DictReader(f))

    hdr = 4
    header_row(
        cov, hdr,
        ["Technique ID", "Sub?", "Name", "Tactics", "Platforms", "In scope (Y/N)",
         "Status (0-3)", "Validation ref.", "Owner", "Required data components"],
        [14, 7, 40, 30, 26, 13, 12, 26, 18, 70],
    )
    dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"0,1,2,3"', allow_blank=True)
    cov.add_data_validation(dv_yn)
    cov.add_data_validation(dv_status)
    for i, t in enumerate(tech):
        rr = hdr + 1 + i
        cov.cell(row=rr, column=1, value=t["technique_id"]).font = Font(name=FONT, size=9, color=ACCENT)
        cov.cell(row=rr, column=2, value=t["is_subtechnique"]).font = Font(name=FONT, size=9, color=INK)
        cov.cell(row=rr, column=3, value=t["name"]).font = Font(name=FONT, size=9, color=INK)
        cov.cell(row=rr, column=4, value=t["tactics"]).font = Font(name=FONT, size=9, color=INK)
        cov.cell(row=rr, column=5, value=t["platforms"]).font = Font(name=FONT, size=9, color=INK)
        for col in (6, 7, 8, 9):
            c = cov.cell(row=rr, column=col)
            c.font = Font(name=FONT, size=9, color="0000FF")
            c.fill = INPUT_FILL
            c.border = BOX
        cov.cell(row=rr, column=10, value=t["data_components"]).font = Font(name=FONT, size=8, color="53616B")
        cov.cell(row=rr, column=10).alignment = Alignment(wrap_text=False)
    first_t, last_t = hdr + 1, hdr + len(tech)
    dv_yn.add(f"F{first_t}:F{last_t}")
    dv_status.add(f"G{first_t}:G{last_t}")
    cov.auto_filter.ref = f"A{hdr}:J{last_t}"
    cov.conditional_formatting.add(
        f"G{first_t}:G{last_t}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F4A6A6",
                       mid_type="num", mid_value=1.5, mid_color="FFE9A8",
                       end_type="num", end_value=3, end_color="A8D5B5"),
    )
    # summary block above the table
    cov["L4"] = "Coverage summary"
    cov["L4"].font = Font(name=FONT, size=11, bold=True, color=NAVY)
    cov.column_dimensions["L"].width = 40
    cov.column_dimensions["M"].width = 14
    summary = [
        ("Techniques in dataset (ATT&CK v%s)" % meta["alignment"]["attack"]["version"], f"=COUNTA(A{first_t}:A{last_t})"),
        ("Techniques in scope", f'=COUNTIF(F{first_t}:F{last_t},"Y")'),
        ("Points scored", f'=SUMIFS(G{first_t}:G{last_t},F{first_t}:F{last_t},"Y")'),
        ("Points available (3 x in scope)", "=3*M6"),
        ("Validated Coverage Score (VCS)", "=IF(M6=0,0,M7/M8)"),
        ("Detection logic or better", '=IF(M6=0,0,COUNTIFS(F%d:F%d,"Y",G%d:G%d,">=2")/M6)' % (first_t, last_t, first_t, last_t)),
        ("Validated by emulation", '=IF(M6=0,0,COUNTIFS(F%d:F%d,"Y",G%d:G%d,"=3")/M6)' % (first_t, last_t, first_t, last_t)),
        ("No telemetry (blind)", '=IF(M6=0,0,COUNTIFS(F%d:F%d,"Y",G%d:G%d,"=0")/M6)' % (first_t, last_t, first_t, last_t)),
    ]
    for i, (label, formula) in enumerate(summary):
        rr = 5 + i
        cov.cell(row=rr, column=12, value=label).font = Font(name=FONT, size=10, color=INK)
        c = cov.cell(row=rr, column=13, value=formula)
        c.font = Font(name=FONT, size=10, bold=(i >= 4), color=NAVY)
        c.number_format = "0.0%" if i >= 4 else "#,##0"
    cov.cell(row=14, column=12, value="A high VCS over a small in-scope set is not coverage. Record the scoping rationale in Setup.").font = Font(name=FONT, size=8, italic=True, color="A03A3A")

    # per-tactic rollup
    cov.cell(row=16, column=12, value="VCS by tactic").font = Font(name=FONT, size=11, bold=True, color=NAVY)
    tactics = sorted({t.strip() for row in tech for t in row["tactics"].split(";") if t.strip()})
    for i, tac in enumerate(tactics):
        rr = 17 + i
        cov.cell(row=rr, column=12, value=tac).font = Font(name=FONT, size=10, color=INK)
        n = f'COUNTIFS(F{first_t}:F{last_t},"Y",D{first_t}:D{last_t},"*{tac}*")'
        p = f'SUMIFS(G{first_t}:G{last_t},F{first_t}:F{last_t},"Y",D{first_t}:D{last_t},"*{tac}*")'
        c = cov.cell(row=rr, column=13, value=f"=IF({n}=0,\"\",{p}/(3*{n}))")
        c.font = Font(name=FONT, size=10, color=NAVY)
        c.number_format = "0.0%"

    # =====================================================================
    # 5. Dashboard
    # =====================================================================
    db = wb.create_sheet("Dashboard")
    db.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHI", (3, 10, 46, 12, 12, 12, 12, 26, 26)):
        db.column_dimensions[col].width = w
    h1(db, "B2", "TID-CMM Dashboard")
    db["B3"] = "=Setup!C4&\"  ·  assessed \"&Setup!C5&\"  ·  \"&Setup!C7"
    db["B3"].font = Font(name=FONT, size=10, italic=True, color=ACCENT)

    header_row(db, 6, ["Domain", "Name", "Weight %", "Score", "Adjusted", "Target", "Gap", "Band", "Constraint applied"],
               [10, 46, 12, 10, 12, 10, 10, 28, 34])
    d_first = 7
    for i, d in enumerate(model.domains):
        rr = d_first + i
        db.cell(row=rr, column=1, value=d.id).font = Font(name=FONT, size=10, bold=True, color=ACCENT)
        db.cell(row=rr, column=2, value=d.name).font = CALC_FONT
        db.cell(row=rr, column=3, value=f"={weights_ref[d.id]}").font = CALC_FONT
        c = db.cell(row=rr, column=4, value=f"={domain_cells[d.id]['score']}")
        c.font = CALC_FONT
        c.number_format = "0.00"
    av_row = d_first + [d.id for d in model.domains].index("AV")
    dc_row = d_first + [d.id for d in model.domains].index("DC")
    de_row = d_first + [d.id for d in model.domains].index("DE")
    for i, d in enumerate(model.domains):
        rr = d_first + i
        if d.id == "AV":
            adj = f"=IF(D{rr}=\"\",\"\",D{rr})"
            note = '=IF(D%d="","","")' % rr
        elif d.id == "DE":
            adj = (f'=IF(D{rr}="","",MIN(D{rr},IF(D{dc_row}="",5,D{dc_row}+1),'
                   f'IF(D{av_row}="",5,D{av_row}+1)))')
            note = (f'=IF(D{rr}="","",IF(D{rr}>MIN(IF(D{dc_row}="",5,D{dc_row}+1),'
                    f'IF(D{av_row}="",5,D{av_row}+1)),IF(D{rr}>IF(D{av_row}="",5,D{av_row}+1),'
                    f'"C1 validation ceiling","C2 visibility ceiling"),""))')
        else:
            adj = f'=IF(D{rr}="","",MIN(D{rr},IF(D{av_row}="",5,D{av_row}+1)))'
            note = f'=IF(D{rr}="","",IF(D{rr}>IF(D{av_row}="",5,D{av_row}+1),"C1 validation ceiling",""))'
        e = db.cell(row=rr, column=5, value=adj)
        e.font = Font(name=FONT, size=10, bold=True, color=NAVY)
        e.number_format = "0.00"
        db.cell(row=rr, column=6, value=f"={targets_ref[d.id]}").font = CALC_FONT
        g = db.cell(row=rr, column=7, value=f'=IF(E{rr}="","",MAX(0,F{rr}-E{rr}))')
        g.font = CALC_FONT
        g.number_format = "0.00"
        db.cell(row=rr, column=8, value=BAND_FORMULA.format(c=f"E{rr}")).font = CALC_FONT
        db.cell(row=rr, column=9, value=note).font = Font(name=FONT, size=9, italic=True, color="A03A3A")
    d_last = d_first + len(model.domains) - 1

    tr = d_last + 1
    db.cell(row=tr, column=2, value="OVERALL (weighted, adjusted)").font = Font(name=FONT, size=12, bold=True, color=NAVY)
    db.cell(row=tr, column=3, value=f"=SUM(C{d_first}:C{d_last})").font = Font(name=FONT, size=10, bold=True)
    raw = db.cell(row=tr, column=4, value=f'=IF(SUM(C{d_first}:C{d_last})=0,"",SUMPRODUCT(C{d_first}:C{d_last},N(D{d_first}:D{d_last}))/SUM(C{d_first}:C{d_last}))')
    raw.font = Font(name=FONT, size=11, color="6B7A85")
    raw.number_format = "0.00"
    ov = db.cell(row=tr, column=5, value=f'=IF(SUM(C{d_first}:C{d_last})=0,"",SUMPRODUCT(C{d_first}:C{d_last},N(E{d_first}:E{d_last}))/SUM(C{d_first}:C{d_last}))')
    ov.font = Font(name=FONT, size=14, bold=True, color=NAVY)
    ov.number_format = "0.00"
    db.cell(row=tr, column=6, value=f'=IF(SUM(C{d_first}:C{d_last})=0,"",SUMPRODUCT(C{d_first}:C{d_last},N(F{d_first}:F{d_last}))/SUM(C{d_first}:C{d_last}))').number_format = "0.00"
    db.cell(row=tr, column=8, value=BAND_FORMULA.format(c=f"E{tr}")).font = Font(name=FONT, size=12, bold=True, color=ACCENT)
    db.cell(row=tr + 1, column=2, value="Column D is the self-assessed score. Column E is after integrity constraints C1/C2. Report column E.").font = Font(name=FONT, size=9, italic=True, color="6B7A85")

    kr = tr + 3
    h2(db, f"B{kr}", "Headline indicators")
    items = [
        ("Overall maturity (adjusted)", f"=IF(E{tr}=\"\",\"\",E{tr})", "0.00"),
        ("Self-assessed before constraints", f"=IF(D{tr}=\"\",\"\",D{tr})", "0.00"),
        ("Weakest domain", f'=IF(COUNT(E{d_first}:E{d_last})=0,"",INDEX(A{d_first}:A{d_last},MATCH(MIN(E{d_first}:E{d_last}),E{d_first}:E{d_last},0)))', "General"),
        ("Strongest domain", f'=IF(COUNT(E{d_first}:E{d_last})=0,"",INDEX(A{d_first}:A{d_last},MATCH(MAX(E{d_first}:E{d_last}),E{d_first}:E{d_last},0)))', "General"),
        ("ATT&CK Validated Coverage Score", "='ATT&CK Coverage'!M9", "0.0%"),
        ("Techniques in scope", "='ATT&CK Coverage'!M6", "#,##0"),
        ("In-scope techniques with no telemetry", "='ATT&CK Coverage'!M12", "0.0%"),
        ("Unevidenced 4s and 5s across all domains", "=" + "+".join(domain_cells[d.id]["challenges"] for d in model.domains), "#,##0"),
        ("Movement since previous assessment", f'=IF(OR(E{tr}="",N(Setup!C12)=0),"n/a",E{tr}-N(Setup!C12))', "+0.00;-0.00;0.00"),
    ]
    for i, (label, formula, fmt) in enumerate(items):
        rr = kr + 1 + i
        db.cell(row=rr, column=2, value=label).font = Font(name=FONT, size=10, color=INK)
        c = db.cell(row=rr, column=4, value=formula)
        c.font = Font(name=FONT, size=11, bold=True, color=NAVY)
        if fmt != "General":
            c.number_format = fmt

    radar = RadarChart()
    radar.type = "marker"
    radar.title = "Maturity by domain (adjusted vs target)"
    radar.style = 26
    data = Reference(db, min_col=5, max_col=6, min_row=6, max_row=d_last)
    cats = Reference(db, min_col=1, min_row=d_first, max_row=d_last)
    radar.add_data(data, titles_from_data=True)
    radar.set_categories(cats)
    radar.y_axis.scaling.min = 0
    radar.y_axis.scaling.max = 5
    radar.height, radar.width = 11, 13
    db.add_chart(radar, "K6")

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Gap to target by domain"
    bar.add_data(Reference(db, min_col=7, max_col=7, min_row=6, max_row=d_last), titles_from_data=True)
    bar.set_categories(cats)
    bar.height, bar.width = 9, 13
    bar.y_axis.title = "Levels"
    db.add_chart(bar, "K30")

    db.conditional_formatting.add(
        f"E{d_first}:E{d_last}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F4A6A6",
                       mid_type="num", mid_value=3, mid_color="FFE9A8",
                       end_type="num", end_value=5, end_color="A8D5B5"),
    )

    # =====================================================================
    # 6. Roadmap
    # =====================================================================
    rm = wb.create_sheet("Roadmap")
    rm.sheet_view.showGridLines = False
    h1(rm, "A1", "Prioritised improvement roadmap")
    rm["A2"] = ("Impact = domain weight x sub-capability weight x gap to target. Sort or filter on Impact "
                "to sequence the work. 'What good looks like next' is the descriptor for the level above "
                "your current score.")
    rm["A2"].font = Font(name=FONT, size=9, italic=True, color="53616B")
    rm.merge_cells("A2:K2")
    header_row(rm, 4, ["ID", "Domain", "Sub-capability", "Current", "Target", "Gap", "Impact",
                       "What good looks like at the next level", "Evidence to produce", "Owner", "Due"],
               [8, 9, 34, 10, 9, 8, 11, 74, 52, 20, 12])
    rr = 5
    for d in model.domains:
        info = domain_cells[d.id]
        for i, sc in enumerate(d.subcapabilities):
            srow = info["first"] + i
            sheet = info["sheet"]
            rm.cell(row=rr, column=1, value=sc.id).font = Font(name=FONT, size=9, bold=True, color=ACCENT)
            rm.cell(row=rr, column=2, value=d.id).font = Font(name=FONT, size=9, color=INK)
            rm.cell(row=rr, column=3, value=sc.name).font = Font(name=FONT, size=9, color=INK)
            rm.cell(row=rr, column=4, value=f"='{sheet}'!E{srow}").font = Font(name=FONT, size=9, color="008000")
            rm.cell(row=rr, column=5, value=f"='{sheet}'!F{srow}").font = Font(name=FONT, size=9, color="008000")
            rm.cell(row=rr, column=6, value=f'=IF(OR(D{rr}="",D{rr}="NA"),"",MAX(0,E{rr}-N(D{rr})))').font = Font(name=FONT, size=9)
            imp = rm.cell(row=rr, column=7, value=f'=IF(F{rr}="","",{weights_ref[d.id]}/100*{sc.weight}/100*F{rr}*1000)')
            imp.font = Font(name=FONT, size=9, bold=True, color=NAVY)
            imp.number_format = "0.0"
            nxt = (f'=IF(OR(D{rr}="",D{rr}="NA"),"",'
                   + "IF(N(D%d)>=5,\"At the top of the scale — sustain and contribute back.\"," % rr
                   + "CHOOSE(N(D%d)+2," % rr
                   + ",".join('"' + sc.levels[lv].replace('"', "'")[:255] + '"' for lv in range(6))
                   + ")))")
            c = rm.cell(row=rr, column=8, value=nxt)
            c.font = Font(name=FONT, size=9, color=INK)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            e = rm.cell(row=rr, column=9, value="; ".join(sc.evidence))
            e.font = Font(name=FONT, size=8, color="53616B")
            e.alignment = Alignment(wrap_text=True, vertical="top")
            rm.cell(row=rr, column=10, value=f"='{sheet}'!I{srow}").font = Font(name=FONT, size=9, color="008000")
            due = rm.cell(row=rr, column=11)
            due.font = INPUT_FONT
            due.fill = INPUT_FILL
            due.border = BOX
            rm.row_dimensions[rr].height = 40
            rr += 1
    rm.auto_filter.ref = f"A4:K{rr-1}"
    rm.conditional_formatting.add(
        f"G5:G{rr-1}",
        ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="F4A6A6"),
    )

    # =====================================================================
    # 7. Crosswalk
    # =====================================================================
    cw = wb.create_sheet("Crosswalk")
    cw.sheet_view.showGridLines = False
    h1(cw, "A1", "Framework crosswalk")
    cw["A2"] = ("Mapping is indicative, at sub-capability level, and intended to let a TID-CMM assessment "
                "feed existing NIST CSF 2.0, SOC-CMM and ISO 27001 reporting without a separate exercise.")
    cw["A2"].font = Font(name=FONT, size=9, italic=True, color="53616B")
    cw.merge_cells("A2:F2")
    header_row(cw, 4, ["ID", "Domain", "Sub-capability", "NIST CSF 2.0", "SOC-CMM", "Other"],
               [9, 9, 40, 40, 40, 34])
    rr = 5
    for d in model.domains:
        for sc in d.subcapabilities:
            cw.cell(row=rr, column=1, value=sc.id).font = Font(name=FONT, size=9, bold=True, color=ACCENT)
            cw.cell(row=rr, column=2, value=d.id).font = Font(name=FONT, size=9)
            cw.cell(row=rr, column=3, value=sc.name).font = Font(name=FONT, size=9)
            cw.cell(row=rr, column=4, value=", ".join(sc.crosswalk.get("nist_csf_2", []))).font = Font(name=FONT, size=9)
            cw.cell(row=rr, column=5, value=", ".join(sc.crosswalk.get("soc_cmm", []))).font = Font(name=FONT, size=9)
            other = []
            for k, v in sc.crosswalk.items():
                if k not in ("nist_csf_2", "soc_cmm"):
                    other.append(f"{k}: {', '.join(v)}")
            cw.cell(row=rr, column=6, value=" | ".join(other)).font = Font(name=FONT, size=9)
            rr += 1
    cw.auto_filter.ref = f"A4:F{rr-1}"

    # =====================================================================
    # 8. Level descriptors reference
    # =====================================================================
    lr = wb.create_sheet("Level descriptors")
    lr.sheet_view.showGridLines = False
    h1(lr, "A1", "Full level descriptors (0-5) for every sub-capability")
    header_row(lr, 3, ["ID", "Domain", "Sub-capability", "Level", "Descriptor"], [9, 9, 34, 8, 130])
    rr = 4
    for d in model.domains:
        for sc in d.subcapabilities:
            for lv in range(6):
                lr.cell(row=rr, column=1, value=sc.id).font = Font(name=FONT, size=9, bold=True, color=ACCENT)
                lr.cell(row=rr, column=2, value=d.id).font = Font(name=FONT, size=9)
                lr.cell(row=rr, column=3, value=sc.name).font = Font(name=FONT, size=9)
                lr.cell(row=rr, column=4, value=lv).font = Font(name=FONT, size=9, bold=True)
                c = lr.cell(row=rr, column=5, value=sc.levels[lv])
                c.font = Font(name=FONT, size=9, color=INK)
                c.alignment = Alignment(wrap_text=True, vertical="top")
                rr += 1
    lr.auto_filter.ref = f"A3:E{rr-1}"

    # Document properties are authored deliberately. Left alone, openpyxl stamps
    # itself as the creator, which is both wrong and untidy.
    wb.properties.creator = "Reza Adineh"
    wb.properties.lastModifiedBy = "Reza Adineh"
    wb.properties.title = f"TID-CMM Self-Assessment v{meta['model']['version']}"
    wb.properties.subject = "Threat-Informed Detection Capability Maturity Model"
    wb.properties.description = (
        "Self-assessment workbook for TID-CMM, aligned to MITRE ATT&CK Enterprise "
        f"v{meta['alignment']['attack']['version']}. https://tid-cmm.com"
    )
    wb.properties.keywords = "detection engineering; MITRE ATT&CK; maturity model; SOC"
    wb.properties.category = "Security assessment"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    _clean_office_metadata(out_path)
    return out_path


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "build" / "TID-CMM-Self-Assessment-v1.0.xlsx"
    print(build(out))
